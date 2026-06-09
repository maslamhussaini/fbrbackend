import openpyxl
from io import BytesIO
from openpyxl.utils import get_column_letter


def _open_workbook(file_bytes: bytes):
    """Open Excel file — handles both .xlsx and .xls formats."""
    # Try xlsx first
    try:
        return openpyxl.load_workbook(BytesIO(file_bytes), data_only=True), None
    except Exception:
        pass
    # Try xls via xlrd
    try:
        import xlrd
        from openpyxl import Workbook
        xls = xlrd.open_workbook(file_contents=file_bytes)
        wb  = Workbook()
        wb.remove(wb.active)
        for sheet_name in xls.sheet_names():
            xls_ws = xls.sheet_by_name(sheet_name)
            ws = wb.create_sheet(title=sheet_name)
            for row in range(xls_ws.nrows):
                for col in range(xls_ws.ncols):
                    cell = xls_ws.cell(row, col)
                    ws.cell(row=row+1, column=col+1, value=cell.value)
        return wb, None
    except Exception as e:
        return None, f"Cannot open file: {str(e)}"

# ── Invoice header columns ────────────────────────────────────────────────────
HEADER_COLUMNS = {
    "invoice date":              "invoice_date",
    "buyer ntn/cnic":            "buyer_ntn_cnic",
    "buyer business name":       "buyer_business_name",
    "buyer province":            "buyer_province",
    "buyer address":             "buyer_address",
    "buyer registration type":   "buyer_registration_type",  # Registered | Unregistered
    "invoice ref no":            "invoice_ref_no",
    "scenario id":               "scenario_id",              # SN001–SN028
}

# ── Item columns ──────────────────────────────────────────────────────────────
ITEM_COLUMNS = {
    "hs code":                          "hs_code",
    "product description":              "product_description",
    "rate":                             "rate",               # "18%" "0%" "Exempt"
    "uom":                              "uom",                # "KG" "Pieces" etc
    "quantity":                         "quantity",
    "value excl st":                    "value_excl_st",
    "sales tax":                        "sales_tax",
    "retail price":                     "retail_price",       # fixedNotifiedValueOrRetailPrice
    "total values":                     "total_values",
    "sale type":                        "sale_type",          # full string
    "sro schedule no":                  "sro_schedule_no",
    "sro item serial no":               "sro_item_serial_no",
    # Optional
    "discount":                         "discount",
    "further tax":                      "further_tax",
    "extra tax":                        "extra_tax",
    "fed payable":                      "fed_payable",
    "st withheld":                      "st_withheld",
}

SCENARIO_DESCRIPTIONS = {
    "SN001": "Registered buyer — standard rate 18%",
    "SN002": "Unregistered buyer — standard rate + further tax 4%",
    "SN005": "Reduced rate (8th Schedule)",
    "SN006": "Exempt goods (6th Schedule)",
    "SN007": "Zero-rated goods",
    "SN008": "3rd Schedule goods",
    "SN016": "Processing/conversion of goods",
    "SN017": "Goods FED in ST mode",
    "SN024": "Goods as per SRO.297(I)/2023",
    "SN028": "Reduced rate unregistered",
}


def parse_excel(file_bytes: bytes) -> dict:
    errors = []
    header = {}
    items = []

    wb, err = _open_workbook(file_bytes)
    if err:
        return {"success": False, "header": {}, "items": [], "errors": [err]}

    # ── Header sheet ──────────────────────────────────────────────────────────
    if "Invoice" not in wb.sheetnames:
        errors.append("Missing sheet 'Invoice' — download the template")
    else:
        ws = wb["Invoice"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            key = str(row[0]).strip().lower()
            if key in HEADER_COLUMNS:
                header[HEADER_COLUMNS[key]] = row[1]

    # ── Items sheet ───────────────────────────────────────────────────────────
    if "Items" not in wb.sheetnames:
        errors.append("Missing sheet 'Items' — download the template")
    else:
        ws = wb["Items"]
        col_map = {}
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                for col_idx, cell in enumerate(row):
                    if cell:
                        key = str(cell).strip().lower()
                        if key in ITEM_COLUMNS:
                            col_map[col_idx] = ITEM_COLUMNS[key]
                continue
            if not any(row):
                continue
            item = {}
            for col_idx, value in enumerate(row):
                if col_idx in col_map:
                    item[col_map[col_idx]] = value
            if item:
                items.append(item)

    return {
        "success": len(errors) == 0,
        "header": header,
        "items": items,
        "errors": errors
    }


def generate_excel_template() -> bytes:
    wb = openpyxl.Workbook()

    # ── Invoice sheet ─────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Invoice"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 40

    ws1["A1"] = "Field"
    ws1["B1"] = "Value"

    header_fields = [
        ("Invoice Date",             "2025-07-12"),
        ("Buyer NTN/CNIC",           "8352312-6"),
        ("Buyer Business Name",      "AADAM TEXTILE"),
        ("Buyer Province",           "SINDH"),
        ("Buyer Address",            "Plot No CR-435, Karachi"),
        ("Buyer Registration Type",  "Registered"),
        ("Invoice Ref No",           "667"),
        ("Scenario ID",              "SN001"),
    ]
    for i, (field, example) in enumerate(header_fields, 2):
        ws1[f"A{i}"] = field
        ws1[f"B{i}"] = example

    # Scenario reference
    ws1["A12"] = "--- Scenario Reference ---"
    for i, (code, desc) in enumerate(SCENARIO_DESCRIPTIONS.items(), 13):
        ws1[f"A{i}"] = code
        ws1[f"B{i}"] = desc

    # ── Items sheet ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Items")
    item_headers = [
        "HS Code", "Product Description", "Rate", "UoM", "Quantity",
        "Value Excl ST", "Sales Tax", "Retail Price", "Total Values",
        "Sale Type", "SRO Schedule No", "SRO Item Serial No",
        "Discount", "Further Tax", "Extra Tax", "FED Payable", "ST Withheld"
    ]
    for col, h in enumerate(item_headers, 1):
        ws2.cell(row=1, column=col, value=h)
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # Example row — SN001 standard goods
    example = [
        "3923.9090", "Plastic Containers", "18%", "KG", 100,
        10000, 1800, 0, 11800,
        "Goods at standard rate (default)", "", "",
        0, 0, 0, 0, 0
    ]
    for col, val in enumerate(example, 1):
        ws2.cell(row=2, column=col, value=val)

    # ── Rate reference sheet ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Rate & SaleType Reference")
    ws3["A1"] = "Scenario"
    ws3["B1"] = "Rate"
    ws3["C1"] = "Sale Type (exact text)"
    ws3["D1"] = "Further Tax"

    ref_data = [
        ("SN001", "18%",    "Goods at standard rate (default)",  "0"),
        ("SN002", "18%",    "Goods at standard rate (default)",  "0.04"),
        ("SN005", "varies", "Goods at Reduced Rate",             "0"),
        ("SN006", "Exempt", "Exempt goods",                      "0"),
        ("SN007", "0%",     "Goods at zero-rate",                "0"),
        ("SN008", "18%",    "3rd Schedule Goods",                "0"),
        ("SN016", "18%",    "Processing/Conversion of Goods",    "0"),
        ("SN017", "17%",    "Goods (FED in ST Mode)",            "0"),
    ]
    for i, row in enumerate(ref_data, 2):
        for j, val in enumerate(row, 1):
            ws3.cell(row=i, column=j, value=val)

    for col in "ABCD":
        ws3.column_dimensions[col].width = 35

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
