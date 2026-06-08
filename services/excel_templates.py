import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


def _header_style(ws, row, headers, fill_color="1a1d27", font_color="4f8ef7"):
    """Apply dark header style."""
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color=font_color, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = max(18, len(h) + 4)


def generate_bulk_template() -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Invoices ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Invoices"
    ws1.row_dimensions[1].height = 28

    inv_headers = [
        "Invoice Date", "Invoice Ref No",
        "Buyer NTN/CNIC", "Buyer Name",
        "Buyer Province", "Buyer Address", "Buyer Registration",
        "HS Code", "Product Description", "Rate", "UoM",
        "Quantity", "Value Excl ST", "Sales Tax",
        "Retail Price", "Further Tax", "Discount",
        "SRO Schedule No", "SRO Item Serial"
    ]
    _header_style(ws1, 1, inv_headers)

    # Example rows
    examples = [
        ["2025-07-12", "INV-001", "8352312-6", "AADAM TEXTILE",
         "SINDH", "KARACHI", "Registered",
         "3923.9090", "Plastic Containers", "18%", "KG",
         100, 10000, 1800, 0, 0, 0, "", ""],
        ["2025-07-12", "INV-002", "0000000", "UNREGISTERED",
         "SINDH", "KARACHI", "Unregistered",
         "6309.0000", "Used Clothing", "18%", "KG",
         800, 124640, 6232, 0, 4966.56, 0, "", ""],
    ]
    for row in examples:
        ws1.append(row)

    # Notes row
    ws1.append([])
    note = ws1.cell(row=4, column=1,
                    value="💡 If Buyer NTN/CNIC matches a saved customer, all buyer fields auto-fill. "
                          "If HS Code matches a saved product, tax fields auto-fill.")
    note.font = Font(color="94a3b8", italic=True, size=10)
    ws1.merge_cells("A4:S4")

    # ── Sheet 2: Customers ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Customers")
    ws2.row_dimensions[1].height = 28

    cust_headers = [
        "Name", "NTN/CNIC", "Registration Status",
        "Province", "City", "Address",
        "WHT %", "Further Tax %"
    ]
    _header_style(ws2, 1, cust_headers, fill_color="0f1117", font_color="22c55e")

    cust_examples = [
        ["AADAM TEXTILE", "8352312-6", "Registered", "SINDH", "KARACHI",
         "PLOT NO CR-435, KARACHI", 0, 0],
        ["RASHID QAMAR", "4240119113683", "Unregistered", "SINDH", "KARACHI",
         "KARACHI", 0, 4],
        ["UNREGISTERED", "0000000", "Unregistered", "SINDH", "KARACHI",
         "PAKISTAN", 0, 4],
    ]
    for row in cust_examples:
        ws2.append(row)

    ws2.append([])
    note2 = ws2.cell(row=5, column=1,
                     value="💡 Registered buyers → SN001 (standard rate). "
                           "Unregistered → SN002 (standard + further tax 4%). Auto-detected.")
    note2.font = Font(color="94a3b8", italic=True, size=10)
    ws2.merge_cells("A5:H5")

    # ── Sheet 3: Products ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Products")
    ws3.row_dimensions[1].height = 28

    prod_headers = [
        "Product Code", "Description", "HS Code",
        "UoM", "Rate", "Sales Tax %",
        "Sale Type", "SRO Schedule No", "SRO Item Serial",
        "FED %", "MRP / Retail Price"
    ]
    _header_style(ws3, 1, prod_headers, fill_color="0f1117", font_color="4f8ef7")

    prod_examples = [
        ["PROD-001", "Plastic Containers", "3923.9090",
         "KG", "18%", 18,
         "Goods at standard rate (default)", "", "", 0, 0],
        ["PROD-002", "Used Clothing", "6309.0000",
         "KG", "18%", 18,
         "Goods at standard rate (default)", "", "", 0, 0],
        ["PROD-003", "Exempt Goods Sample", "0101.2100",
         "KG", "Exempt", 0,
         "Exempt goods", "6th Schd Table I", "100", 0, 0],
    ]
    for row in prod_examples:
        ws3.append(row)

    # ── Sheet 4: Reference ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Reference")
    _header_style(ws4, 1, ["Scenario", "Description", "Rate", "Sale Type", "When to use"],
                  fill_color="0a0d14", font_color="94a3b8")

    ref_data = [
        ["SN001", "Registered buyer standard rate", "18%",
         "Goods at standard rate (default)", "Most common"],
        ["SN002", "Unregistered buyer", "18%",
         "Goods at standard rate (default)", "Walk-in / unregistered + 4% further tax"],
        ["SN005", "Reduced rate 8th Schedule", "varies",
         "Goods at Reduced Rate", "Fill SRO Schedule No + SRO Item Serial"],
        ["SN006", "Exempt goods", "Exempt",
         "Exempt goods", "Fill SRO Schedule No + SRO Item Serial"],
        ["SN007", "Zero-rated", "0%",
         "Goods at zero-rate", "Exports etc."],
        ["SN008", "3rd Schedule", "18%",
         "3rd Schedule Goods", "Tax on retail price, not value"],
        ["SN016", "Processing/Conversion", "18%",
         "Processing/Conversion of Goods", ""],
        ["SN017", "FED in ST mode", "17%",
         "Goods (FED in ST Mode)", ""],
    ]
    for row in ref_data:
        ws4.append(row)
    for col in range(1, 6):
        ws4.column_dimensions[get_column_letter(col)].width = 30

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
